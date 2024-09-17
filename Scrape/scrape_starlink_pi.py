import requests
from bs4 import BeautifulSoup
import re
from urllib.parse import urlparse
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

DRIVER_PATH = "/usr/bin/chromedriver"  # Change to your path

def append_to_excel(filename, data):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Link", "City", "Country", "Download Speed", "Upload Speed", "Latency"])

    for row in data:
        ws.append(row)

    wb.save(filename)

def get_airbnb_listings(api_key, cx, search_country, start=1, num=10):
    base_url = "https://www.googleapis.com/customsearch/v1"
    query = f"starlink {search_country} site:https://airbnb.com inurl:airbnb.com -inurl:stays -inurl:villas -inurl:houses -inurl:pet-friendly"
    params = {
        "key": api_key,
        "cx": cx,
        "q": query,
        "start": start,
        "num": num
    }

    response = requests.get(base_url, params=params)
    if response.status_code == 200:
        data = response.json()
        airbnb_links = [item['link'] for item in data.get('items', [])]
        return airbnb_links, start + num  # Increment start by num
    else:
        print(f"Failed to fetch Airbnb listings for {search_country}:", response.text)
        return [], start + num  # Increment start by num even in case of failure

def extract_airbnb_id(url):
    url_components = urlparse(url)
    path_components = url_components.path.split('/')
    airbnb_id = path_components[-1] if path_components[-1] else path_components[-2]
    return airbnb_id

def extract_location(html_content):
    soup = BeautifulSoup(html_content, "html.parser")
    div_tag = soup.find("div", class_="_16e70jgn")
    if div_tag:
        c1yo0219_div = div_tag.find("div", class_="c1yo0219")
        if c1yo0219_div:
            toieuka_div = c1yo0219_div.find("div", class_="toieuka")
            if toieuka_div:
                h2_tag = toieuka_div.find("h2", class_="hpipapi")
                if h2_tag:
                    location_text = h2_tag.text.strip()
                    city, country = extract_city_country(location_text)
                    return city, country
    return None, None

def extract_city_country(location_text):
    city_country_match = re.search(r'in (.+?), (.+)', location_text)
    if city_country_match:
        city = city_country_match.group(1)
        country = city_country_match.group(2)
        return city, country
    return None, None
def main(api_key, cx, countries):
    for search_country in countries:
        processed_links = set()
        rows = []

        # Load existing data from the Excel file to avoid duplicates
        try:
            wb = load_workbook("airbnb_starlink_speeds.xlsx")
            ws = wb.active
            existing_urls = set([cell.value for cell in ws["A"]])
            processed_links.update(existing_urls)
        except FileNotFoundError:
            pass

        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
        options.add_argument(f'user-agent={user_agent}')
        service = Service(executable_path=DRIVER_PATH)

        with webdriver.Chrome(service=service, options=options) as driver:
            start_index = 1
            while True:
                airbnb_links, start_index = get_airbnb_listings(api_key, cx, search_country, start_index)
                if not airbnb_links:
                    print(f"All listings processed for {search_country}")
                    break

                for link in airbnb_links:
                    if link in processed_links:
                        print(f"Skipping already processed link: {link}")
                        continue

                    try:
                        airbnb_id = extract_airbnb_id(link)

                        airbnb_url = f"https://www.airbnb.com/rooms/{airbnb_id}"
                        print("Processing:", airbnb_url)

                        driver.get(airbnb_url)
                        time.sleep(3)
                        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        time.sleep(3)
                        html_content = driver.page_source
                        city, country = extract_location(html_content)

                        if city and country:
                            row = [airbnb_url, city, country]
                            rows.append(row)
                            append_to_excel("airbnb_starlink_speeds_new.xlsx", [row])
                            print("Speed written for " + airbnb_url)
                            processed_links.add(link)
                        else:
                            print(f"Location format not recognized for {airbnb_url}")

                    except Exception as e:
                        print(f"Error processing {link}: {e}")
                        print(f"Skipping {airbnb_url} due to processing error")

                print("Moving to the next page with start index:", start_index)

if __name__ == "__main__":
    API_KEY = "AIzaSyAOjGMLisl1Jp8USLsQdJbFJNHhCOGJidY"
    CX = "34c31d66702ef400b"
    #countries = ["Argentina", "Indonesia", "Thailand", "Philippines", "Brazil", "Vietnam"]
    #countries = ["guatemala","costa rica", "chile", "ecuador", "japan", "uruguay"]
    #countries = ["uruguay", "peru","canada", "norway", "kenya", "greece"]
    #countries = ["argentina", "new zealand", "iceland", "malaysia", "japan"]
    countries = ["indonesia", "spain", "france", "puerto rico", "jamaica"]
    main(API_KEY, CX, countries)
